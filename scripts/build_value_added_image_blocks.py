#!/usr/bin/env python3
"""Extract value-added service images and build supplemental image blocks.

This is stage 2 after `build_value_added_text_blocks.py`.

The script does not OCR images. It extracts/de-duplicates embedded workbook
images, records source provenance and nearby cell text, and optionally merges
human/LLM visual annotations from `manual_image_annotations.json` into
retrievable image supplement blocks.

Default output:
  sop/structured/value_added_images/
    images/
    contact_sheets/
    image_inventory.json
    image_blocks.jsonl
    manifest.json
    manual_image_annotations.json  # optional input, created externally

The summary workbook `增值业务大全20260105 (2).xlsx` is skipped by default.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import re
import shutil
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

import openpyxl
from PIL import Image, ImageDraw, ImageFont

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from scripts.build_value_added_text_blocks import (  # noqa: E402
    SUMMARY_WORKBOOK,
    VALUE_ADDED_FOLDER,
    clean_text,
    infer_intent_tags,
    service_for_path,
    stable_id,
)

DEFAULT_OUTPUT_DIR = ROOT / "sop" / "structured" / "value_added_images"
ANNOTATIONS_FILE = "manual_image_annotations.json"


def iter_sources(include_summary_workbook: bool) -> list[Path]:
    sources: list[Path] = []
    if VALUE_ADDED_FOLDER.exists():
        for path in sorted(VALUE_ADDED_FOLDER.glob("*.xlsx")):
            if not include_summary_workbook and path.name == SUMMARY_WORKBOOK:
                continue
            sources.append(path)
    return sources


def image_ext(image: Any) -> str:
    fmt = str(getattr(image, "format", "") or "").lower()
    if fmt in {"jpeg", "jpg"}:
        return "jpg"
    if fmt in {"png", "gif", "bmp"}:
        return fmt
    return "png"


def image_anchor(image: Any) -> tuple[int | None, int | None]:
    try:
        return image.anchor._from.row + 1, image.anchor._from.col + 1
    except Exception:
        return None, None


def nearby_text(ws: Any, row: int | None, col: int | None, radius_rows: int = 3, radius_cols: int = 3) -> str:
    if row is None or col is None:
        return ""
    lines: list[str] = []
    r1 = max(1, row - radius_rows)
    r2 = min(ws.max_row, row + radius_rows)
    c1 = max(1, col - radius_cols)
    c2 = min(ws.max_column, col + radius_cols)
    for r in range(r1, r2 + 1):
        cells: list[str] = []
        for c in range(c1, c2 + 1):
            text = clean_text(ws.cell(r, c).value)
            if text:
                cells.append(text)
        if cells:
            lines.append(f"R{r}: " + " | ".join(cells))
    return "\n".join(lines)


def build_inventory(output_dir: Path, include_summary_workbook: bool) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    images_dir = output_dir / "images"
    images_dir.mkdir(parents=True, exist_ok=True)

    inventory: list[dict[str, Any]] = []
    errors: list[dict[str, str]] = []
    unique_by_hash: dict[str, dict[str, Any]] = {}
    next_id = 1

    for source_path in iter_sources(include_summary_workbook):
        service = service_for_path(source_path)
        try:
            workbook = openpyxl.load_workbook(source_path, read_only=False, data_only=True)
        except Exception as exc:
            errors.append({"source_file": str(source_path.relative_to(ROOT)), "error": repr(exc)})
            continue

        for ws in workbook.worksheets:
            for image_idx, image in enumerate(getattr(ws, "_images", []) or [], start=1):
                try:
                    data = image._data()
                except Exception as exc:
                    errors.append({
                        "source_file": str(source_path.relative_to(ROOT)),
                        "sheet": ws.title,
                        "image_idx": str(image_idx),
                        "error": repr(exc),
                    })
                    continue
                digest = hashlib.sha256(data).hexdigest()
                if digest not in unique_by_hash:
                    image_id = f"img_{next_id:04d}"
                    next_id += 1
                    ext = image_ext(image)
                    image_path = images_dir / f"{image_id}_{digest[:12]}.{ext}"
                    image_path.write_bytes(data)
                    try:
                        with Image.open(image_path) as pil_image:
                            actual_w, actual_h = pil_image.size
                    except Exception:
                        actual_w = getattr(image, "width", None)
                        actual_h = getattr(image, "height", None)
                    unique_by_hash[digest] = {
                        "image_id": image_id,
                        "hash": digest,
                        "image_path": str(image_path.relative_to(ROOT)),
                        "actual_width": actual_w,
                        "actual_height": actual_h,
                        "references": [],
                    }

                row, col = image_anchor(image)
                rel_source = str(source_path.relative_to(ROOT))
                reference = {
                    "source_file": rel_source,
                    "sheet": ws.title,
                    "image_idx": image_idx,
                    "anchor": {"row": row, "col": col},
                    "declared_width": getattr(image, "width", None),
                    "declared_height": getattr(image, "height", None),
                    "nearby_text": nearby_text(ws, row, col),
                }
                unique_by_hash[digest]["references"].append(reference)

    for item in unique_by_hash.values():
        first_ref = item["references"][0] if item["references"] else {}
        source_path = ROOT / first_ref.get("source_file", "")
        service = service_for_path(source_path) if source_path.exists() else None
        all_nearby = "\n".join(ref.get("nearby_text", "") for ref in item["references"] if ref.get("nearby_text"))
        inventory.append({
            "image_id": item["image_id"],
            "hash": item["hash"],
            "service_id": service.service_id if service else None,
            "service_name": service.canonical_name if service else None,
            "image_path": item["image_path"],
            "actual_width": item["actual_width"],
            "actual_height": item["actual_height"],
            "references": item["references"],
            "primary_source_file": first_ref.get("source_file"),
            "primary_sheet": first_ref.get("sheet"),
            "primary_anchor": first_ref.get("anchor"),
            "nearby_text": all_nearby[:2400],
            "intent_tags_from_nearby": infer_intent_tags(all_nearby),
        })

    inventory.sort(key=lambda r: int(re.search(r"\d+", r["image_id"]).group(0)))
    return inventory, errors


def wrapped_text(draw: ImageDraw.ImageDraw, text: str, font: ImageFont.ImageFont, max_width: int) -> list[str]:
    words = re.split(r"([/ _\\-])", text)
    lines: list[str] = []
    line = ""
    for word in words:
        candidate = line + word
        if draw.textbbox((0, 0), candidate, font=font)[2] <= max_width or not line:
            line = candidate
        else:
            lines.append(line)
            line = word
    if line:
        lines.append(line)
    return lines[:5]


def make_contact_sheets(output_dir: Path, inventory: list[dict[str, Any]], *, per_page: int = 6) -> list[str]:
    contact_dir = output_dir / "contact_sheets"
    if contact_dir.exists():
        shutil.rmtree(contact_dir)
    contact_dir.mkdir(parents=True, exist_ok=True)

    try:
        font = ImageFont.truetype("/System/Library/Fonts/PingFang.ttc", 15)
        title_font = ImageFont.truetype("/System/Library/Fonts/PingFang.ttc", 18)
    except Exception:
        font = ImageFont.load_default()
        title_font = ImageFont.load_default()

    cols = 2
    thumb_w = 520
    thumb_h = 430
    label_h = 110
    rows = math.ceil(per_page / cols)
    page_paths: list[str] = []

    for page_idx, start in enumerate(range(0, len(inventory), per_page), start=1):
        page_items = inventory[start:start + per_page]
        canvas = Image.new("RGB", (cols * thumb_w, rows * (thumb_h + label_h)), "white")
        draw = ImageDraw.Draw(canvas)
        for offset, item in enumerate(page_items):
            x = (offset % cols) * thumb_w
            y = (offset // cols) * (thumb_h + label_h)
            image_path = ROOT / item["image_path"]
            try:
                img = Image.open(image_path).convert("RGB")
                img.thumbnail((thumb_w - 20, thumb_h - 20))
                canvas.paste(img, (x + (thumb_w - img.width) // 2, y + (thumb_h - img.height) // 2))
            except Exception as exc:
                draw.text((x + 10, y + 10), f"IMAGE ERROR: {exc}", fill="red", font=font)

            label_y = y + thumb_h
            draw.rectangle([x, label_y, x + thumb_w, label_y + label_h], fill=(244, 244, 244), outline=(210, 210, 210))
            source_name = Path(item.get("primary_source_file") or "").name
            anchor = item.get("primary_anchor") or {}
            label = (
                f"{item['image_id']} | {item.get('service_name') or '-'} | "
                f"{source_name} / {item.get('primary_sheet') or '-'} / "
                f"R{anchor.get('row')}C{anchor.get('col')}"
            )
            draw.text((x + 8, label_y + 6), label, fill="black", font=title_font)
            nearby = (item.get("nearby_text") or "").replace("\n", " ")
            for line_idx, line in enumerate(wrapped_text(draw, nearby[:240], font, thumb_w - 16)):
                draw.text((x + 8, label_y + 34 + line_idx * 16), line, fill=(35, 35, 35), font=font)
        page_path = contact_dir / f"contact_{page_idx:02d}.jpg"
        canvas.save(page_path, quality=92)
        page_paths.append(str(page_path.relative_to(ROOT)))
    return page_paths


def load_annotations(output_dir: Path) -> dict[str, dict[str, Any]]:
    path = output_dir / ANNOTATIONS_FILE
    if not path.exists():
        return {}
    data = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(data, dict) and "annotations" in data:
        items = data["annotations"]
    elif isinstance(data, list):
        items = data
    else:
        raise ValueError(f"unsupported annotation format: {path}")
    return {str(item["image_id"]): item for item in items if item.get("image_id")}


def build_blocks(inventory: list[dict[str, Any]], annotations: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    blocks: list[dict[str, Any]] = []
    for item in inventory:
        ann = annotations.get(item["image_id"], {})
        image_type = ann.get("image_type") or "unknown"
        importance = ann.get("importance") or "unreviewed"
        visual_summary = ann.get("visual_summary") or ""
        key_information = ann.get("key_information") or []
        if isinstance(key_information, str):
            key_information = [key_information]
        prompt_usage = ann.get("prompt_usage") or ""
        suggested_intent_tags = ann.get("intent_tags") or item.get("intent_tags_from_nearby") or []
        if isinstance(suggested_intent_tags, str):
            suggested_intent_tags = [suggested_intent_tags]

        extracted_text = "\n".join(
            part
            for part in [
                f"图片类型：{image_type}",
                f"重要性：{importance}",
                f"视觉摘要：{visual_summary}" if visual_summary else "",
                "关键信息：\n" + "\n".join(f"- {x}" for x in key_information) if key_information else "",
                f"使用建议：{prompt_usage}" if prompt_usage else "",
                "相邻表格文本：\n" + item.get("nearby_text", "") if item.get("nearby_text") else "",
            ]
            if part
        )
        source = item.get("primary_source_file") or ""
        block_id = f"value_added_image:{stable_id(item['image_id'], source, item.get('hash', '')[:12])}"
        display_source = source
        if item.get("primary_sheet"):
            display_source += f" / {item['primary_sheet']}"
        anchor = item.get("primary_anchor") or {}
        if anchor.get("row") is not None:
            display_source += f" / R{anchor.get('row')}C{anchor.get('col')}"
        embedding_text = "\n".join(
            part
            for part in [
                f"服务：{item.get('service_name')}",
                f"图片补充：{visual_summary}" if visual_summary else "",
                "关键信息：" + "；".join(key_information) if key_information else "",
                item.get("nearby_text", ""),
            ]
            if part
        )
        blocks.append({
            "block_id": block_id,
            "image_id": item["image_id"],
            "service_id": item.get("service_id"),
            "service_name": item.get("service_name"),
            "domain": "活动",
            "category": "增值服务",
            "modality": "image_manual_visual",
            "source_type": "xlsx_image",
            "source_file": source,
            "sheet": item.get("primary_sheet"),
            "anchor": item.get("primary_anchor"),
            "image_path": item.get("image_path"),
            "image_type": image_type,
            "importance": importance,
            "intent_tags": suggested_intent_tags,
            "visual_summary": visual_summary,
            "key_information": key_information,
            "prompt_usage": prompt_usage,
            "nearby_text": item.get("nearby_text", ""),
            "text": extracted_text,
            "embedding_text": embedding_text,
            "display_text": f"【服务】{item.get('service_name')}\n【来源】{display_source}\n{extracted_text}",
            "references": item.get("references", []),
            "annotation_status": "reviewed" if ann else "unreviewed",
        })
    return blocks


def write_jsonl(path: Path, records: list[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8") as f:
        for record in records:
            f.write(json.dumps(record, ensure_ascii=False) + "\n")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR))
    parser.add_argument("--include-summary-workbook", action="store_true")
    parser.add_argument("--per-page", type=int, default=6)
    args = parser.parse_args()

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    inventory, errors = build_inventory(output_dir, args.include_summary_workbook)
    contact_sheets = make_contact_sheets(output_dir, inventory, per_page=args.per_page)
    annotations = load_annotations(output_dir)
    blocks = build_blocks(inventory, annotations)

    (output_dir / "image_inventory.json").write_text(json.dumps({"images": inventory}, ensure_ascii=False, indent=2), encoding="utf-8")
    write_jsonl(output_dir / "image_blocks.jsonl", blocks)

    status_counts = Counter(block["annotation_status"] for block in blocks)
    importance_counts = Counter(block["importance"] for block in blocks)
    type_counts = Counter(block["image_type"] for block in blocks)
    service_counts = Counter(block["service_id"] for block in blocks)
    manifest = {
        "scope": "value_added_images",
        "summary_workbook_skipped": not args.include_summary_workbook,
        "summary_workbook_name": SUMMARY_WORKBOOK,
        "image_count": len(inventory),
        "image_block_count": len(blocks),
        "annotation_count": len(annotations),
        "annotation_status_counts": dict(status_counts),
        "importance_counts": dict(importance_counts),
        "image_type_counts": dict(type_counts),
        "images_by_service": dict(sorted(service_counts.items())),
        "contact_sheets": contact_sheets,
        "errors": errors,
        "outputs": {
            "inventory": "image_inventory.json",
            "blocks": "image_blocks.jsonl",
            "manual_annotations": ANNOTATIONS_FILE,
            "images_dir": "images/",
            "contact_sheets_dir": "contact_sheets/",
        },
    }
    (output_dir / "manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"images: {len(inventory)}")
    print(f"annotations: {len(annotations)}")
    print(f"contact sheets: {len(contact_sheets)}")
    print(f"output: {output_dir.relative_to(ROOT)}")
    if errors:
        print(f"errors: {len(errors)}", file=sys.stderr)


if __name__ == "__main__":
    main()
