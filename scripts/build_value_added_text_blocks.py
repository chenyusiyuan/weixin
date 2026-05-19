#!/usr/bin/env python3
"""Build weakly-structured text blocks for value-added service SOP assets.

This stage deliberately does *not* force the new xlsx/docx assets into one
business schema. The source files mix key-value sheets, comparison matrices,
QA tables, free-form notes, and embedded screenshots. The output therefore
keeps source-local structure and provenance, so a later retrieval stage can
match by service aliases and feed only relevant text blocks to the prompt.

Default inputs:
  - sop/债务咨询顾问服务费（还款无忧）场景（12.11更新）
  - sop/增值服务大全（未上翻）

Default output:
  - sop/structured/value_added_text/services.json
  - sop/structured/value_added_text/text_blocks.jsonl
  - sop/structured/value_added_text/manifest.json

The summary workbook `增值业务大全20260105 (2).xlsx` is skipped by default
because it duplicates the single-service files and would inflate retrieval.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
import zipfile
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

import openpyxl


ROOT = Path(__file__).resolve().parents[1]

DEBT_FOLDER = ROOT / "sop" / "债务咨询顾问服务费（还款无忧）场景（12.11更新）"
VALUE_ADDED_FOLDER = ROOT / "sop" / "增值服务大全（未上翻）"
DEFAULT_OUTPUT_DIR = ROOT / "sop" / "structured" / "value_added_text"
SUMMARY_WORKBOOK = "增值业务大全20260105 (2).xlsx"

PUNCT_SPLIT_RE = re.compile(r"[、/，,（）()【】\[\]《》<>\\s]+")
SPACE_RE = re.compile(r"\\s+")
QUESTION_HEADER_RE = re.compile(r"^(问题\\d+|\\d+[.．、]|步骤[一二三四五六七八九十]+|客服流程图|对客话术QA)")


@dataclass
class ServiceSpec:
    service_id: str
    canonical_name: str
    aliases: set[str] = field(default_factory=set)
    source_names: set[str] = field(default_factory=set)

    def to_json(self) -> dict[str, Any]:
        return {
            "service_id": self.service_id,
            "canonical_name": self.canonical_name,
            "aliases": sorted(a for a in self.aliases if a),
            "source_names": sorted(self.source_names),
        }


SERVICE_SPECS: dict[str, ServiceSpec] = {
    "zhonghui_insurance": ServiceSpec(
        "zhonghui_insurance",
        "众惠财产相互保险社",
        {"众惠财产相互保险社", "众惠保险", "众惠财产保险", "众惠"},
    ),
    "premium_card": ServiceSpec(
        "premium_card",
        "优享卡",
        {"优享卡", "优享卡服务", "优享卡权益", "优享卡费用"},
    ),
    "member_benefits": ServiceSpec(
        "member_benefits",
        "会员活动权益",
        {"会员活动", "卡豆会员权益", "会员权益", "会员权益明细"},
    ),
    "member_pay_later": ServiceSpec(
        "member_pay_later",
        "会员（先享后付）",
        {"会员", "先享后付", "会员先享后付", "尊享VIP", "VIP"},
    ),
    "credit_risk_check": ServiceSpec(
        "credit_risk_check",
        "信用风险/贷前必查/借钱必查活动",
        {"信用风险", "贷前必查", "借钱必查", "信用风险活动", "天创信用"},
    ),
    "loan_coupon_campaign": ServiceSpec(
        "loan_coupon_campaign",
        "借款优惠券活动",
        {"借款优惠券", "优惠券活动", "借款券", "息费券"},
    ),
    "loan_treasure": ServiceSpec(
        "loan_treasure",
        "借款宝",
        {"借款宝", "维信卡卡贷借款宝"},
    ),
    "debt_consulting_service_fee": ServiceSpec(
        "debt_consulting_service_fee",
        "债务咨询顾问服务费/还款无忧",
        {"债务咨询顾问服务费", "债务咨询顾问服务", "还款无忧", "还款无忧服务", "还款无忧服务费"},
    ),
    "acceleration_card": ServiceSpec(
        "acceleration_card",
        "加速卡",
        {"加速卡", "加速放款", "优先放款"},
    ),
    "merged_loan_campaign": ServiceSpec(
        "merged_loan_campaign",
        "合并借款活动",
        {"合并借款", "合并借款活动"},
    ),
    "interest_fee_deduction_coupon": ServiceSpec(
        "interest_fee_deduction_coupon",
        "息费抵扣金",
        {"息费抵扣金", "抵扣金", "息费优惠券"},
    ),
    "claw_egg_lottery_campaign": ServiceSpec(
        "claw_egg_lottery_campaign",
        "抓娃娃/砸金蛋抽奖活动",
        {"抓娃娃", "砸金蛋", "抽奖活动"},
    ),
    "rejection_cashback": ServiceSpec(
        "rejection_cashback",
        "拒就返",
        {"拒就返", "拒就返权益", "全部被拒返现"},
    ),
    "quota_increase_campaign": ServiceSpec(
        "quota_increase_campaign",
        "提额活动",
        {"提额活动", "公积金认证提额", "学历认证提额", "提额"},
    ),
    "new_year_deduction_campaign": ServiceSpec(
        "new_year_deduction_campaign",
        "新年送抵扣金活动",
        {"新年送抵扣金", "抵扣金活动"},
    ),
    "benefits_monthly_card": ServiceSpec(
        "benefits_monthly_card",
        "权益月卡",
        {"权益月卡", "新品体验月卡", "新品体验权益"},
    ),
    "telemarketing_campaign": ServiceSpec(
        "telemarketing_campaign",
        "电销活动",
        {"电销活动", "电销", "电话营销活动"},
    ),
    "phone_credit_coupon": ServiceSpec(
        "phone_credit_coupon",
        "省话费/话费券包/领红包",
        {"省话费", "话费券包", "领红包", "话费券", "话费充值"},
    ),
    "doudou_score": ServiceSpec(
        "doudou_score",
        "豆豆分",
        {"豆豆分", "荣誉等级", "信用分"},
    ),
    "enforceable_notary": ServiceSpec(
        "enforceable_notary",
        "赋强公证/法信赋强公证",
        {"赋强公证", "法信赋强公证", "赋强公证协议", "诚信计划"},
    ),
    "light_card": ServiceSpec(
        "light_card",
        "轻享卡",
        {"轻享卡", "轻享卡续费", "轻享卡扣费"},
    ),
    "referral_cash_daily_money": ServiceSpec(
        "referral_cash_daily_money",
        "邀好友得现金/天天领钱",
        {"邀好友得现金", "天天领钱", "邀请好友", "好友奖励"},
    ),
    "referral_bounty": ServiceSpec(
        "referral_bounty",
        "邀请赚赏金",
        {"邀请赚赏金", "邀请赏金", "邀请奖励"},
    ),
    "collect_cards_free_order": ServiceSpec(
        "collect_cards_free_order",
        "集卡赢免单活动",
        {"集卡赢免单", "集卡", "免单活动"},
    ),
    "reserved_loan_campaign": ServiceSpec(
        "reserved_loan_campaign",
        "预约借款活动",
        {"预约借款", "预约订单", "预约审核", "预约放款"},
    ),
}

FILENAME_TO_SERVICE_ID = {
    "众惠财产相互保险社": "zhonghui_insurance",
    "优享卡": "premium_card",
    "会员活动-卡豆会员权益明细": "member_benefits",
    "会员（先享后付）": "member_pay_later",
    "信用风险、贷前必查、借钱必查活动": "credit_risk_check",
    "借款优惠券活动": "loan_coupon_campaign",
    "借款宝-维信卡卡贷": "loan_treasure",
    "债务咨询顾问服务费": "debt_consulting_service_fee",
    "债务咨询顾问服务费（还款无忧）问题流程梳理": "debt_consulting_service_fee",
    "还款无忧服务？": "debt_consulting_service_fee",
    "还款无忧服务？？": "debt_consulting_service_fee",
    "加速卡": "acceleration_card",
    "合并借款活动": "merged_loan_campaign",
    "息费抵扣金": "interest_fee_deduction_coupon",
    "抓娃娃、砸金蛋抽奖活动": "claw_egg_lottery_campaign",
    "拒就返（豆豆钱）": "rejection_cashback",
    "提额活动": "quota_increase_campaign",
    "新年送抵扣金活动": "new_year_deduction_campaign",
    "权益月卡": "benefits_monthly_card",
    "电销活动": "telemarketing_campaign",
    "省话费、话费券包、领红包": "phone_credit_coupon",
    "豆豆分": "doudou_score",
    "赋强公证协议、法信赋强公证": "enforceable_notary",
    "轻享卡": "light_card",
    "邀好友得现金、天天领钱": "referral_cash_daily_money",
    "邀请赚赏金": "referral_bounty",
    "集卡赢免单活动": "collect_cards_free_order",
    "预约借款活动": "reserved_loan_campaign",
}

INTENT_KEYWORDS = {
    "inquiry": ["是什么", "含义", "介绍", "权益", "规则", "有什么", "说明", "活动介绍"],
    "cancel": ["取消", "关闭", "不需要", "退订", "续费管理"],
    "refund": ["退费", "退款", "退钱", "返现", "原路退回"],
    "bill_repayment": ["扣款", "还款", "账单", "期数", "逾期", "征信", "银行卡", "待还"],
    "operation_path": ["路径", "入口", "APP", "客服系统", "页面", "按钮", "查看"],
    "escalation": ["工单", "升级", "主管", "回电", "作业", "反馈"],
    "contact": ["客服热线", "热线", "电话", "邮箱", "服务商"],
}


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\u3000", " ")
    text = SPACE_RE.sub(" ", text).strip()
    return text


def stable_id(*parts: str) -> str:
    raw = "|".join(parts)
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()[:10]


def split_aliases(name: str) -> set[str]:
    aliases = {name.strip()}
    for part in PUNCT_SPLIT_RE.split(name):
        part = part.strip()
        if len(part) >= 2:
            aliases.add(part)
    return aliases


def service_for_path(path: Path) -> ServiceSpec:
    stem = path.stem.strip()
    service_id = FILENAME_TO_SERVICE_ID.get(stem)
    if service_id is None:
        service_id = "service_" + stable_id(stem)
        SERVICE_SPECS.setdefault(service_id, ServiceSpec(service_id, stem, {stem}))
    spec = SERVICE_SPECS[service_id]
    spec.source_names.add(stem)
    spec.aliases.update(split_aliases(stem))
    return spec


def infer_intent_tags(text: str) -> list[str]:
    tags = []
    for tag, kws in INTENT_KEYWORDS.items():
        if any(kw in text for kw in kws):
            tags.append(tag)
    return tags


def make_block(
    *,
    service: ServiceSpec,
    source_path: Path,
    source_type: str,
    text: str,
    block_type: str,
    title: str = "",
    sheet: str | None = None,
    row_start: int | None = None,
    row_end: int | None = None,
    extra: dict[str, Any] | None = None,
) -> dict[str, Any] | None:
    text = text.strip()
    if not text:
        return None
    rel_source = str(source_path.relative_to(ROOT))
    source_bits = [rel_source, sheet or "", str(row_start or ""), str(row_end or ""), title, text[:80]]
    block_id = f"value_added_text:{stable_id(*source_bits)}"
    aliases = sorted(a for a in service.aliases if a and a in text)
    if not aliases:
        aliases = sorted(service.aliases)[:6]

    row_range = None
    if row_start is not None and row_end is not None:
        row_range = f"{row_start}-{row_end}" if row_start != row_end else str(row_start)

    embedding_text = "\n".join(
        part
        for part in [
            f"服务：{service.canonical_name}",
            f"别名：{'、'.join(sorted(service.aliases)[:12])}",
            f"来源：{rel_source}" + (f" / {sheet}" if sheet else ""),
            f"标题：{title}" if title else "",
            text,
        ]
        if part
    )
    display_text = "\n".join(
        part
        for part in [
            f"【服务】{service.canonical_name}",
            f"【来源】{rel_source}" + (f" / {sheet}" if sheet else "") + (f" / 行 {row_range}" if row_range else ""),
            f"【标题】{title}" if title else "",
            text,
        ]
        if part
    )

    record = {
        "block_id": block_id,
        "service_id": service.service_id,
        "service_name": service.canonical_name,
        "service_aliases": aliases,
        "domain": "活动",
        "category": "增值服务",
        "source_type": source_type,
        "source_file": rel_source,
        "sheet": sheet,
        "row_range": row_range,
        "block_type": block_type,
        "title": title,
        "intent_tags": infer_intent_tags(text),
        "text": text,
        "embedding_text": embedding_text,
        "display_text": display_text,
    }
    if extra:
        record.update(extra)
    return record


def non_empty_row(ws: Any, row_idx: int) -> list[str]:
    values = []
    for col_idx in range(1, ws.max_column + 1):
        value = clean_text(ws.cell(row_idx, col_idx).value)
        values.append(value)
    return values


def find_header_row(rows: list[tuple[int, list[str]]]) -> tuple[int, dict[str, int]] | None:
    for row_idx, values in rows[:20]:
        joined = "|".join(values[:8])
        if "步骤" in joined and ("用户问句" in joined or "应对话术" in joined):
            mapping: dict[str, int] = {}
            for i, value in enumerate(values):
                if value in {"步骤", "用户问句", "应对话术", "注意事项", "解决方案"}:
                    mapping[value] = i
            return row_idx, mapping
    return None


def extract_qa_sheet_blocks(service: ServiceSpec, source_path: Path, ws: Any, rows: list[tuple[int, list[str]]]) -> list[dict[str, Any]]:
    header = find_header_row(rows)
    if not header:
        return []
    header_row_idx, col = header
    blocks: list[dict[str, Any]] = []
    last_step = ""
    last_query = ""
    last_notes = ""
    for row_idx, values in rows:
        if row_idx <= header_row_idx:
            continue
        step = values[col.get("步骤", 0)] if "步骤" in col else ""
        query = values[col.get("用户问句", 1)] if "用户问句" in col else ""
        script = values[col.get("应对话术", 2)] if "应对话术" in col else ""
        notes = values[col.get("注意事项", 3)] if "注意事项" in col else ""
        solution = values[col.get("解决方案", 4)] if "解决方案" in col else ""
        if step:
            last_step = step
        if query:
            last_query = query
        if notes:
            last_notes = notes

        fields = {
            "步骤": step or last_step,
            "用户问句": query or last_query,
            "应对话术": script,
            "注意事项": notes or last_notes,
            "解决方案": solution,
        }
        if not any(fields.values()):
            continue
        text = "\n".join(f"{k}：{v}" for k, v in fields.items() if v)
        block = make_block(
            service=service,
            source_path=source_path,
            source_type="xlsx",
            text=text,
            block_type="qa_row",
            title=fields.get("步骤", ""),
            sheet=ws.title,
            row_start=row_idx,
            row_end=row_idx,
            extra={"user_query": fields.get("用户问句", ""), "script": fields.get("应对话术", "")},
        )
        if block:
            blocks.append(block)
    return blocks


def row_to_line(row_idx: int, values: list[str]) -> str:
    cells = [value for value in values if value]
    if not cells:
        return ""
    return f"R{row_idx}: " + " | ".join(cells)


def extract_general_sheet_blocks(
    service: ServiceSpec,
    source_path: Path,
    ws: Any,
    rows: list[tuple[int, list[str]]],
    *,
    max_chars: int,
    max_rows: int,
) -> list[dict[str, Any]]:
    blocks: list[dict[str, Any]] = []
    current: list[tuple[int, str]] = []

    def flush() -> None:
        nonlocal current
        if not current:
            return
        row_start = current[0][0]
        row_end = current[-1][0]
        text = "\n".join(line for _, line in current)
        title = ""
        if current:
            title = current[0][1].split("|", 1)[0].replace(f"R{row_start}:", "").strip()[:80]
        block = make_block(
            service=service,
            source_path=source_path,
            source_type="xlsx",
            text=text,
            block_type="sheet_rows",
            title=title,
            sheet=ws.title,
            row_start=row_start,
            row_end=row_end,
        )
        if block:
            blocks.append(block)
        current = []

    previous_row_idx = None
    for row_idx, values in rows:
        line = row_to_line(row_idx, values)
        if not line:
            flush()
            previous_row_idx = None
            continue
        if previous_row_idx is not None and row_idx - previous_row_idx > 1:
            flush()
        current_text_len = sum(len(line_item) for _, line_item in current)
        if current and (len(current) >= max_rows or current_text_len + len(line) > max_chars):
            flush()
        current.append((row_idx, line))
        previous_row_idx = row_idx
    flush()
    return blocks


def extract_xlsx_blocks(source_path: Path, *, max_chars: int, max_rows: int) -> list[dict[str, Any]]:
    service = service_for_path(source_path)
    workbook = openpyxl.load_workbook(source_path, read_only=False, data_only=True)
    blocks: list[dict[str, Any]] = []
    for ws in workbook.worksheets:
        rows: list[tuple[int, list[str]]] = []
        for row_idx in range(1, ws.max_row + 1):
            values = non_empty_row(ws, row_idx)
            if any(values):
                rows.append((row_idx, values))
        if not rows:
            continue

        qa_blocks = extract_qa_sheet_blocks(service, source_path, ws, rows)
        if qa_blocks:
            blocks.extend(qa_blocks)
        else:
            blocks.extend(
                extract_general_sheet_blocks(
                    service,
                    source_path,
                    ws,
                    rows,
                    max_chars=max_chars,
                    max_rows=max_rows,
                )
            )
    return blocks


def docx_paragraphs(source_path: Path) -> list[str]:
    with zipfile.ZipFile(source_path) as zf:
        xml = zf.read("word/document.xml")
    root = ET.fromstring(xml)
    paragraphs = []
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    for para in root.findall(".//w:p", ns):
        parts: list[str] = []
        for node in para.iter():
            tag = node.tag.rsplit("}", 1)[-1]
            if tag == "t" and node.text:
                parts.append(node.text)
            elif tag == "tab":
                parts.append(" ")
        text = clean_text("".join(parts))
        if text:
            paragraphs.append(text)
    return paragraphs


def extract_docx_blocks(source_path: Path, *, max_chars: int) -> list[dict[str, Any]]:
    service = service_for_path(source_path)
    paragraphs = docx_paragraphs(source_path)
    blocks: list[dict[str, Any]] = []
    current: list[tuple[int, str]] = []
    current_title = ""

    def flush() -> None:
        nonlocal current, current_title
        if not current:
            return
        para_start = current[0][0]
        para_end = current[-1][0]
        text = "\n".join(t for _, t in current)
        block = make_block(
            service=service,
            source_path=source_path,
            source_type="docx",
            text=text,
            block_type="docx_paragraphs",
            title=current_title,
            row_start=para_start,
            row_end=para_end,
            extra={"paragraph_range": f"{para_start}-{para_end}" if para_start != para_end else str(para_start)},
        )
        if block:
            blocks.append(block)
        current = []
        current_title = ""

    for idx, text in enumerate(paragraphs, start=1):
        is_heading = bool(QUESTION_HEADER_RE.match(text)) or text.endswith("问题")
        current_len = sum(len(t) for _, t in current)
        if current and (is_heading or current_len + len(text) > max_chars):
            flush()
        if not current_title and (is_heading or len(text) < 40):
            current_title = text[:120]
        current.append((idx, text))
    flush()
    return blocks


def iter_sources(include_summary_workbook: bool) -> list[Path]:
    sources: list[Path] = []
    if DEBT_FOLDER.exists():
        sources.extend(sorted(DEBT_FOLDER.glob("*.xlsx")))
        sources.extend(sorted(DEBT_FOLDER.glob("*.docx")))
    if VALUE_ADDED_FOLDER.exists():
        for path in sorted(VALUE_ADDED_FOLDER.glob("*.xlsx")):
            if not include_summary_workbook and path.name == SUMMARY_WORKBOOK:
                continue
            sources.append(path)
    return sources


def write_jsonl(path: Path, records: list[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8") as f:
        for record in records:
            f.write(json.dumps(record, ensure_ascii=False) + "\n")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR))
    parser.add_argument("--include-summary-workbook", action="store_true")
    parser.add_argument("--max-chars", type=int, default=1800)
    parser.add_argument("--max-rows", type=int, default=10)
    args = parser.parse_args()

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    records: list[dict[str, Any]] = []
    errors: list[dict[str, str]] = []
    no_text_sources: list[str] = []
    sources = iter_sources(args.include_summary_workbook)
    for source_path in sources:
        try:
            if source_path.suffix.lower() == ".xlsx":
                blocks = extract_xlsx_blocks(source_path, max_chars=args.max_chars, max_rows=args.max_rows)
            elif source_path.suffix.lower() == ".docx":
                blocks = extract_docx_blocks(source_path, max_chars=args.max_chars)
            else:
                continue
            records.extend(blocks)
            if not blocks:
                service_for_path(source_path)
                no_text_sources.append(str(source_path.relative_to(ROOT)))
        except Exception as exc:  # pragma: no cover - diagnostic output
            errors.append({"source_file": str(source_path.relative_to(ROOT)), "error": repr(exc)})

    records.sort(key=lambda r: (r["service_id"], r["source_file"], r.get("sheet") or "", r.get("row_range") or ""))

    services = [spec.to_json() for spec in sorted(SERVICE_SPECS.values(), key=lambda s: s.service_id) if spec.source_names]
    service_counts = Counter(r["service_id"] for r in records)
    source_counts = Counter(r["source_file"] for r in records)
    intent_counts = Counter(tag for r in records for tag in r.get("intent_tags", []))

    write_jsonl(output_dir / "text_blocks.jsonl", records)
    (output_dir / "services.json").write_text(
        json.dumps({"services": services}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    manifest = {
        "scope": "value_added_text",
        "summary_workbook_skipped": not args.include_summary_workbook,
        "summary_workbook_name": SUMMARY_WORKBOOK,
        "source_count": len(sources),
        "service_count": len(services),
        "text_block_count": len(records),
        "errors": errors,
        "no_text_sources": no_text_sources,
        "blocks_by_service": dict(sorted(service_counts.items())),
        "blocks_by_source": dict(sorted(source_counts.items())),
        "intent_tag_counts": dict(sorted(intent_counts.items())),
        "outputs": {
            "services": "services.json",
            "text_blocks": "text_blocks.jsonl",
        },
    }
    (output_dir / "manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    print(f"sources: {len(sources)}")
    print(f"services: {len(services)}")
    print(f"text blocks: {len(records)}")
    print(f"output: {output_dir.relative_to(ROOT)}")
    if errors:
        print(f"errors: {len(errors)}", file=sys.stderr)
        for error in errors[:10]:
            print(json.dumps(error, ensure_ascii=False), file=sys.stderr)


if __name__ == "__main__":
    main()
